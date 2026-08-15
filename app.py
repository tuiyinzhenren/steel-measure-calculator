import streamlit as st
import pandas as pd
import math
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime

# 页面配置
st.set_page_config(
    page_title="钢结构吊装机械措施费计算工具",
    page_icon="🏗️",
    layout="wide"
)

# 初始化session_state
if 'buildings' not in st.session_state:
    st.session_state.buildings = []
if 'equipment' not in st.session_state:
    st.session_state.equipment = []
if 'aux_times' not in st.session_state:
    st.session_state.aux_times = []
if 'materials' not in st.session_state:
    st.session_state.materials = []
if 'calc_results' not in st.session_state:
    st.session_state.calc_results = None

# 定额数据
QUOTA_DATA = {
    "2015": {
        "汽车吊_120t以下": 12, "汽车吊_120-300t": 9, "汽车吊_300t以上": 5,
        "履带吊_50-130t": 13, "履带吊_150-300t": 9, "履带吊_300t以上_常规": 7, "履带吊_300t以上_超起": 4,
        "塔吊_50m以下": 16, "塔吊_50-150m": 12, "塔吊_150-250m": 10, "塔吊_250m以上": 9
    },
    "2021": {
        "汽车吊_100t以下": 12, "汽车吊_100-300t": 9, "汽车吊_300t以上": 5,
        "履带吊_50-150t": 13, "履带吊_150-300t": 9, "履带吊_300t以上_常规": 7, "履带吊_300t以上_超起": 4,
        "塔吊_50m以下": 16, "塔吊_50-150m": 12, "塔吊_150-250m": 10, "塔吊_250m以上": 9
    }
}

ADJUSTMENT_COEFFICIENTS = {
    "单层厂房": {"实腹式柱": 1.2, "格构式柱_小": 0.75, "格构式柱_大": 0.4, "主梁/吊车梁": 1.0, "次构件": 2.0},
    "多层结构": {"实腹式柱": 1.0, "倾斜柱": 0.75, "主梁": 1.2, "次梁": 1.8},
    "高层/超高层": {"垂直柱": 1.0, "倾斜柱": 0.75, "主梁": 1.2, "次梁": 1.8, "支撑": 0.75, "伸臂/环桁架散件": 0.6},
    "空间网架/网壳": {"杆件": 2.5, "檩条": 3.5, "马道": 0.25, "分块吊装": 1.0, "补杆": 1.5},
    "空间桁架": {"地面拼装": 2.0, "高空拼装": 1.5, "主桁架": 0.3, "次桁架": 0.6},
    "异形特殊结构": {"异形构件": 0.3}
}

COMPONENT_OPTIONS = {
    "单层厂房": ["实腹式柱", "格构式柱", "主梁/吊车梁", "次构件"],
    "多层结构": ["实腹式柱", "倾斜柱", "主梁", "次梁"],
    "高层/超高层": ["垂直柱", "倾斜柱", "主梁", "次梁", "支撑", "伸臂/环桁架散件"],
    "空间网架/网壳": ["杆件", "檩条", "马道", "分块吊装", "补杆"],
    "空间桁架": ["地面拼装", "高空拼装", "主桁架", "次桁架"],
    "异形特殊结构": ["异形构件"]
}

EQUIPMENT_SPECS = {
    "2015": {
        "汽车吊": ["120t以下", "120-300t", "300t以上"],
        "履带吊": ["50-130t", "150-300t", "300t以上_常规", "300t以上_超起"],
        "塔吊": ["50m以下", "50-150m", "150-250m", "250m以上"],
        "曲臂车": ["自定义"]
    },
    "2021": {
        "汽车吊": ["100t以下", "100-300t", "300t以上"],
        "履带吊": ["50-150t", "150-300t", "300t以上_常规", "300t以上_超起"],
        "塔吊": ["50m以下", "50-150m", "150-250m", "250m以上"],
        "曲臂车": ["自定义"]
    }
}

STRUCTURE_TYPES = ["单层厂房", "多层结构", "高层/超高层", "空间网架/网壳", "空间桁架", "异形特殊结构"]

# 标题
st.title("🏗️ 钢结构吊装机械措施费计算工具")

# ==================== 一、项目基本信息 ====================
st.header("一、项目基本信息")
col1, col2, col3, col4, col5 = st.columns(5)
with col1:
    project_name = st.text_input("工程名称", value="")
with col2:
    quota_version = st.selectbox("定额版本", ["2015版（浙江精工）", "2021版（精工工业建筑·试行）"], index=1)
    quota_version = "2015" if "2015" in quota_version else "2021"
with col3:
    contract_days = st.number_input("合同工期（天）", value=120, min_value=0)
with col4:
    work_days_per_month = st.number_input("每月工作天数", value=25, min_value=1)
with col5:
    steel_weight = st.number_input("用钢量（吨）", value=0, min_value=0)

col6, col7, col8 = st.columns(3)
with col6:
    author = st.text_input("编制人", value="")
with col7:
    reviewer = st.text_input("审核人", value="")
with col8:
    doc_date = st.date_input("日期", value=datetime.date.today())

# ==================== 二、单体结构信息 ====================
st.header("二、单体结构信息")
st.caption("为每个单体定义结构类型，后续机械措施会自动关联")

# 添加单体按钮
if st.button("+ 添加单体", key="add_building"):
    st.session_state.buildings.append({
        "name": "",
        "structure_type": "",
        "height": 0
    })

# 显示单体列表
for i, bld in enumerate(st.session_state.buildings):
    cols = st.columns([1, 3, 3, 2, 1])
    with cols[0]:
        st.text(f"单体 {i+1}")
    with cols[1]:
        bld["name"] = st.text_input("单体名称", value=bld["name"], key=f"bld_name_{i}")
    with cols[2]:
        bld["structure_type"] = st.selectbox("结构类型", [""] + STRUCTURE_TYPES, 
                                              index=STRUCTURE_TYPES.index(bld["structure_type"]) + 1 if bld["structure_type"] in STRUCTURE_TYPES else 0,
                                              key=f"bld_type_{i}")
    with cols[3]:
        bld["height"] = st.number_input("结构高度(m)", value=bld["height"], min_value=0, key=f"bld_height_{i}")
    with cols[4]:
        if st.button("删除", key=f"del_bld_{i}"):
            st.session_state.buildings.pop(i)
            st.rerun()

# ==================== 三、机械措施 ====================
st.header("三、机械措施")
st.caption("录入设备和构件信息，系统自动查表计算。调整系数可在定额值基础上微调（±20%以内），也可选择“自定义”手动输入效率值。")

# 添加设备配置按钮
if st.button("+ 添加设备配置", key="add_equipment"):
    st.session_state.equipment.append({
        "building_idx": "",
        "eq_type": "",
        "eq_spec": "",
        "eq_qty": 1,
        "component_type": "",
        "pieces": 0,
        "weight": 0.0,
        "base_efficiency": 0,
        "adjust_coeff": 1.0,
        "actual_efficiency": 0,
        "required_shifts": 0,
        "required_days": 0,
        "mobilize_times": 1,
        "shift_freq": 1,
        "remark": ""
    })

# 显示设备列表
for i, eq in enumerate(st.session_state.equipment):
    with st.container():
        # 第一行：基本信息
        cols = st.columns([1, 2, 2, 2, 1, 2, 1])
        with cols[0]:
            st.text(f"序号 {i+1}")
        with cols[1]:
            bld_options = [""] + [b["name"] for b in st.session_state.buildings if b["name"]]
            eq["building_idx"] = st.selectbox("单体名称", bld_options,
                                              index=bld_options.index(eq["building_idx"]) if eq["building_idx"] in bld_options else 0,
                                              key=f"eq_bld_{i}")
        with cols[2]:
            eq["eq_type"] = st.selectbox("设备类型", ["", "汽车吊", "履带吊", "塔吊", "曲臂车/其他"],
                                         index=["", "汽车吊", "履带吊", "塔吊", "曲臂车/其他"].index(eq["eq_type"]) if eq["eq_type"] in ["", "汽车吊", "履带吊", "塔吊", "曲臂车/其他"] else 0,
                                         key=f"eq_type_{i}")
        with cols[3]:
            eq_type_clean = eq["eq_type"].replace("/其他", "")
            spec_options = [""] + EQUIPMENT_SPECS.get(quota_version, {}).get(eq_type_clean, [])
            eq["eq_spec"] = st.selectbox("设备规格", spec_options,
                                        index=spec_options.index(eq["eq_spec"]) if eq["eq_spec"] in spec_options else 0,
                                        key=f"eq_spec_{i}")
        with cols[4]:
            eq["eq_qty"] = st.number_input("数量(台)", value=eq["eq_qty"], min_value=1, key=f"eq_qty_{i}")
        with cols[5]:
            # 根据单体获取结构类型
            structure_type = ""
            bld_height = 0
            if eq["building_idx"]:
                for bld in st.session_state.buildings:
                    if bld["name"] == eq["building_idx"]:
                        structure_type = bld["structure_type"]
                        bld_height = bld["height"]
                        break
            comp_options = ["自定义"] + (COMPONENT_OPTIONS.get(structure_type, []) if structure_type else [])
            eq["component_type"] = st.selectbox("构件类型", [""] + comp_options,
                                               index=comp_options.index(eq["component_type"]) + 1 if eq["component_type"] in comp_options else 0,
                                               key=f"eq_comp_{i}")
        with cols[6]:
            if st.button("删除", key=f"del_eq_{i}"):
                st.session_state.equipment.pop(i)
                st.rerun()
        
        # 第二行：计算参数
        cols2 = st.columns([2, 2, 2, 2, 2, 2, 2])
        with cols2[0]:
            eq["pieces"] = st.number_input("构件数量(件)", value=eq["pieces"], min_value=0, key=f"eq_pieces_{i}")
        with cols2[1]:
            eq["weight"] = st.number_input("单件重量(吨)", value=eq["weight"], min_value=0.0, step=0.1, key=f"eq_weight_{i}")
        with cols2[2]:
            # 计算基准效率
            eq_type_clean = eq["eq_type"].replace("/其他", "")
            quota_key = f"{eq_type_clean}_{eq['eq_spec']}"
            base_eff = QUOTA_DATA.get(quota_version, {}).get(quota_key, 0)
            eq["base_efficiency"] = base_eff
            st.caption("基准效率")
            st.markdown(f"**{base_eff if base_eff else '—'}**")
        with cols2[3]:
            # 调整系数
            standard_coeff = 1.0
            if eq["component_type"] and eq["component_type"] != "自定义" and structure_type:
                standard_coeff = ADJUSTMENT_COEFFICIENTS.get(structure_type, {}).get(eq["component_type"], 1.0)
            
            # 格构式柱双条件自动判断：≤5t 或高度<20m → 小柱(0.75)；>5t且高度≥20m → 大柱(0.4)
            if eq["component_type"] == "格构式柱" and structure_type == "单层厂房":
                standard_coeff = 0.75 if (eq["weight"] <= 5 or bld_height < 20) else 0.4
            
            coeff = st.number_input("调整系数", value=eq["adjust_coeff"], step=0.01, key=f"eq_coeff_{i}")
            eq["adjust_coeff"] = coeff
            
            # 验证提示
            if abs(coeff - standard_coeff) / standard_coeff > 0.2:
                st.warning(f"⚠️ 超出标准值{standard_coeff}的±20%范围")
        
        with cols2[4]:
            # 实际效率
            if eq["component_type"] == "自定义":
                actual_eff = coeff
            else:
                actual_eff = eq["base_efficiency"] * coeff
            eq["actual_efficiency"] = actual_eff
            st.caption("实际效率")
            st.markdown(f"**{actual_eff:.1f}**" if actual_eff > 0 else "**—**")
        with cols2[5]:
            # 计算台班
            if eq["pieces"] > 0 and actual_eff > 0:
                shifts = math.ceil(eq["pieces"] / actual_eff)
                days = math.ceil(shifts / eq["eq_qty"])
            else:
                shifts = 0
                days = 0
            eq["required_shifts"] = shifts
            eq["required_days"] = days
            st.caption("总台班 / 所需天数")
            st.markdown(f"**{shifts} / {days}**")
        with cols2[6]:
            eq["mobilize_times"] = st.number_input("进出场次数", value=eq["mobilize_times"], min_value=0, key=f"eq_mob_{i}")
            eq["shift_freq"] = st.number_input("移位次数/吊", value=eq["shift_freq"], min_value=1, key=f"eq_shift_freq_{i}")
        
        # 第三行：备注
        eq["remark"] = st.text_input("备注", value=eq["remark"], key=f"eq_remark_{i}")
        
        st.divider()

# ==================== 四、辅助时间配置 ====================
st.header("四、辅助时间配置")
st.caption("根据定额规定，以下辅助工作需额外占用天数，计入总工期")

if st.button("+ 添加辅助时间项", key="add_aux"):
    st.session_state.aux_times.append({
        "work": "",
        "eq_info": "",
        "times": 1,
        "per": 1.0,
        "total": 0,
        "remark": ""
    })

for i, aux in enumerate(st.session_state.aux_times):
    cols = st.columns([3, 2, 1, 1, 1, 2, 1])
    with cols[0]:
        aux["work"] = st.selectbox("辅助工作内容", ["", "塔吊爬升", "内爬塔吊拆除", "履带吊安拆", "履带吊工况转换", "路基箱配重倒运", "设备进出场占用", "其他"],
                                   index=["", "塔吊爬升", "内爬塔吊拆除", "履带吊安拆", "履带吊工况转换", "路基箱配重倒运", "设备进出场占用", "其他"].index(aux["work"]) if aux["work"] in ["", "塔吊爬升", "内爬塔吊拆除", "履带吊安拆", "履带吊工况转换", "路基箱配重倒运", "设备进出场占用", "其他"] else 0,
                                   key=f"aux_work_{i}")
    with cols[1]:
        aux["eq_info"] = st.text_input("适用设备", value=aux["eq_info"], key=f"aux_eq_{i}")
    with cols[2]:
        aux["times"] = st.number_input("次数", value=aux["times"], min_value=0, key=f"aux_times_{i}")
    with cols[3]:
        aux["per"] = st.number_input("每次天数", value=aux["per"], min_value=0.0, step=0.5, key=f"aux_per_{i}")
    with cols[4]:
        aux_total = aux["times"] * aux["per"]
        aux["total"] = aux_total
        st.caption("合计天数")
        st.markdown(f"**{aux_total:g}**")
    with cols[5]:
        aux["remark"] = st.text_input("备注", value=aux["remark"], key=f"aux_remark_{i}")
    with cols[6]:
        if st.button("删除", key=f"del_aux_{i}"):
            st.session_state.aux_times.pop(i)
            st.rerun()

# ==================== 五、费用单价 ====================
st.header("五、费用单价")
st.caption("填写各类设备和辅助的单价，用于汇总费用")

with st.expander("设备台班单价", expanded=True):
    cols = st.columns(4)
    with cols[0]:
        price_qiche = st.number_input("汽车吊台班单价（元/台班）", value=0, min_value=0, key="price_qiche")
    with cols[1]:
        price_lvdai = st.number_input("履带吊台班单价（元/台班）", value=0, min_value=0, key="price_lvdai")
    with cols[2]:
        price_tadiao = st.number_input("塔吊台班单价（元/台班）", value=0, min_value=0, key="price_tadiao")
    with cols[3]:
        price_other = st.number_input("曲臂车/其他（元/台班）", value=0, min_value=0, key="price_other")

with st.expander("进出场费单价", expanded=True):
    cols = st.columns(4)
    with cols[0]:
        fee_qiche = st.number_input("汽车吊进出场费（元/次）", value=0, min_value=0, key="fee_qiche")
    with cols[1]:
        fee_lvdai = st.number_input("履带吊进出场费（元/次）", value=0, min_value=0, key="fee_lvdai")
    with cols[2]:
        fee_tadiao = st.number_input("塔吊进出场费（元/次）", value=0, min_value=0, key="fee_tadiao")
    with cols[3]:
        fee_other = st.number_input("其他设备进出场费（元/次）", value=0, min_value=0, key="fee_other")

with st.expander("辅助机械费", expanded=True):
    cols = st.columns(2)
    with cols[0]:
        fee_aux_machine = st.number_input("路基箱/配重/轨道倒运（元/天）", value=0, min_value=0, key="fee_aux_machine")
    with cols[1]:
        fee_aux_labor = st.number_input("辅助人工费（元/天）", value=0, min_value=0, key="fee_aux_labor")

# ==================== 六、材料措施 ====================
st.header("六、材料措施")

if st.button("+ 添加材料措施", key="add_material"):
    st.session_state.materials.append({
        "building_idx": "",
        "content": "",
        "unit": "吨",
        "qty": 0,
        "days": 0,
        "remark": ""
    })

for i, mat in enumerate(st.session_state.materials):
    cols = st.columns([1, 2, 3, 1, 1, 2, 2, 1])
    with cols[0]:
        st.text(f"序号 {i+1}")
    with cols[1]:
        bld_options = [""] + [b["name"] for b in st.session_state.buildings if b["name"]]
        mat["building_idx"] = st.selectbox("单体名称", bld_options,
                                          index=bld_options.index(mat["building_idx"]) if mat["building_idx"] in bld_options else 0,
                                          key=f"mat_bld_{i}")
    with cols[2]:
        mat["content"] = st.text_input("措施内容", value=mat["content"], key=f"mat_content_{i}")
    with cols[3]:
        mat["unit"] = st.text_input("单位", value=mat["unit"], key=f"mat_unit_{i}")
    with cols[4]:
        mat["qty"] = st.number_input("数量", value=mat["qty"], min_value=0, key=f"mat_qty_{i}")
    with cols[5]:
        mat["days"] = st.number_input("使用时间(天)", value=mat["days"], min_value=0, key=f"mat_days_{i}")
    with cols[6]:
        mat["remark"] = st.text_input("备注", value=mat["remark"], key=f"mat_remark_{i}")
    with cols[7]:
        if st.button("删除", key=f"del_mat_{i}"):
            st.session_state.materials.pop(i)
            st.rerun()

# ==================== 七、备注事项 ====================
st.header("七、备注事项")
remarks = st.text_area("填写施工方法说明、特殊要求、现场条件等", value="", height=150)

# ==================== 计算按钮 ====================
st.markdown("---")
col_calc, col_export = st.columns(2)
with col_calc:
    calc_clicked = st.button("🧮 计算", type="primary", width="stretch")
with col_export:
    export_clicked = st.button("📊 导出Excel", width="stretch")

# ==================== 八、计算结果 ====================
if calc_clicked:
    with st.spinner("正在计算..."):
        # 汇总数据
        total_shifts = 0
        total_days = 0
        by_type = {}
        by_bld = {}
        total_mobilize_cost = 0
        
        for eq in st.session_state.equipment:
            shifts = eq["required_shifts"]
            days = eq["required_days"]
            eq_type = eq["eq_type"].replace("/其他", "")
            
            total_shifts += shifts
            total_days += days
            
            # 按设备类型汇总
            if eq_type not in by_type:
                by_type[eq_type] = {"shifts": 0, "days": 0, "mobilize": 0, "count": 0}
            by_type[eq_type]["shifts"] += shifts
            by_type[eq_type]["days"] += days
            by_type[eq_type]["mobilize"] += eq["mobilize_times"]
            by_type[eq_type]["count"] += 1
            
            # 按单体汇总
            if eq["building_idx"]:
                if eq["building_idx"] not in by_bld:
                    by_bld[eq["building_idx"]] = {"shifts": 0, "days": 0}
                by_bld[eq["building_idx"]]["shifts"] += shifts
                by_bld[eq["building_idx"]]["days"] += days
            
            # 进出场费（按设备类型映射到对应单价变量）
            fee_map = {"汽车吊": fee_qiche, "履带吊": fee_lvdai, "塔吊": fee_tadiao, "曲臂车": fee_other}
            mobilize_fee = fee_map.get(eq_type, fee_other)
            total_mobilize_cost += eq["mobilize_times"] * mobilize_fee
        
        # 辅助时间
        total_aux_days = sum(aux["total"] for aux in st.session_state.aux_times)
        grand_total_days = total_days + total_aux_days
        months = grand_total_days / work_days_per_month if work_days_per_month > 0 else 0
        
        # 费用计算
        price_map = {
            "汽车吊": price_qiche,
            "履带吊": price_lvdai,
            "塔吊": price_tadiao,
            "曲臂车": price_other
        }
        
        total_shift_cost = 0
        cost_by_type = {}
        for eq_type, data in by_type.items():
            unit_price = price_map.get(eq_type, price_other)
            cost = data["shifts"] * unit_price
            total_shift_cost += cost
            cost_by_type[eq_type] = {"shifts": data["shifts"], "unit_price": unit_price, "cost": cost}
        
        aux_other_cost = (fee_aux_machine + fee_aux_labor) * total_aux_days
        total_cost = total_shift_cost + total_mobilize_cost + aux_other_cost
        cost_per_ton = total_cost / steel_weight if steel_weight > 0 else 0
        
        # 保存结果
        st.session_state.calc_results = {
            "total_shifts": total_shifts,
            "total_days": total_days,
            "total_aux_days": total_aux_days,
            "grand_total_days": grand_total_days,
            "months": months,
            "total_cost": total_cost,
            "cost_per_ton": cost_per_ton,
            "by_type": by_type,
            "cost_by_type": cost_by_type,
            "by_bld": by_bld,
            "total_mobilize_cost": total_mobilize_cost,
            "total_shift_cost": total_shift_cost,
            "aux_other_cost": aux_other_cost
        }

# 显示结果
if st.session_state.calc_results:
    st.header("八、计算结果")
    r = st.session_state.calc_results
    
    # 概览卡片
    col1, col2, col3, col4, col5 = st.columns(5)
    with col1:
        st.metric("总吊装台班", f"{r['total_shifts']} 台班")
    with col2:
        st.metric("纯吊装天数", f"{r['total_days']} 天")
    with col3:
        st.metric("辅助占用天数", f"{r['total_aux_days']} 天")
    with col4:
        st.metric("总需求天数", f"{r['grand_total_days']} 天", f"约{r['months']:.1f}月")
    with col5:
        diff = contract_days - r['grand_total_days']
        if diff >= 0:
            st.success(f"✓ 工期富余 {diff} 天")
        else:
            st.error(f"⚠ 工期不足，超出 {abs(diff)} 天")
    
    st.metric("措施费合计", f"¥{r['total_cost']:,.0f}", 
              f"折合 ¥{r['cost_per_ton']:,.2f}/吨" if steel_weight > 0 else "")
    
    # 按设备类型汇总
    st.subheader("按设备类型汇总")
    type_data = []
    for eq_type, data in r['by_type'].items():
        cost_info = r['cost_by_type'][eq_type]
        fee_map = {"汽车吊": fee_qiche, "履带吊": fee_lvdai, "塔吊": fee_tadiao, "曲臂车": fee_other}
        mobilize_unit_fee = fee_map.get(eq_type, fee_other)
        m_cost = data['mobilize'] * mobilize_unit_fee
        
        type_data.append({
            "设备类型": eq_type,
            "总台班": data['shifts'],
            "总台数": data['count'],
            "进出场次": data['mobilize'],
            "台班单价(元)": cost_info['unit_price'],
            "台班费(元)": cost_info['cost'],
            "进出场费(元)": m_cost,
            "小计(元)": cost_info['cost'] + m_cost
        })
    
    if r['aux_other_cost'] > 0:
        type_data.append({
            "设备类型": "辅助机械/人工",
            "总台班": None,
            "总台数": None,
            "进出场次": None,
            "台班单价(元)": fee_aux_machine + fee_aux_labor,
            "台班费(元)": None,
            "进出场费(元)": None,
            "小计(元)": r['aux_other_cost']
        })
    
    # 合计行
    sum_shift_fee = sum(d['cost'] for d in r['cost_by_type'].values())
    sum_mobilize_fee = r['total_mobilize_cost']
    type_data.append({
        "设备类型": "合计",
        "总台班": r['total_shifts'],
        "总台数": None,
        "进出场次": None,
        "台班单价(元)": None,
        "台班费(元)": sum_shift_fee,
        "进出场费(元)": sum_mobilize_fee,
        "小计(元)": r['total_cost']
    })
    
    st.dataframe(pd.DataFrame(type_data), width="stretch", hide_index=True)
    
    # 按单体汇总
    if r['by_bld']:
        st.subheader("按单体汇总")
        bld_data = []
        for bld_name, data in r['by_bld'].items():
            bld_data.append({
                "单体名称": bld_name,
                "总台班": data['shifts'],
                "吊装天数": data['days']
            })
        bld_data.append({
            "单体名称": "合计",
            "总台班": r['total_shifts'],
            "吊装天数": r['total_days']
        })
        st.dataframe(pd.DataFrame(bld_data), width="stretch", hide_index=True)
    
    # 辅助时间明细
    if st.session_state.aux_times:
        st.subheader("辅助时间明细")
        aux_data = []
        for aux in st.session_state.aux_times:
            aux_data.append({
                "辅助工作": aux['work'],
                "适用设备": aux['eq_info'],
                "次数": aux['times'],
                "每次天数": aux['per'],
                "合计天数": aux['total'],
                "备注": aux['remark']
            })
        aux_data.append({
            "辅助工作": "合计",
            "适用设备": None,
            "次数": None,
            "每次天数": None,
            "合计天数": r['total_aux_days'],
            "备注": None
        })
        st.dataframe(pd.DataFrame(aux_data), width="stretch", hide_index=True)
    
    # 费用明细
    st.subheader("费用明细")
    fee_data = [
        {"费用项目": "设备台班费", "金额（元）": r['total_shift_cost']},
        {"费用项目": "设备进出场费", "金额（元）": r['total_mobilize_cost']},
        {"费用项目": "辅助机械/人工费", "金额（元）": r['aux_other_cost']},
        {"费用项目": "措施费合计", "金额（元）": r['total_cost']}
    ]
    if steel_weight > 0:
        fee_data.append({"费用项目": "折合单价（元/吨）", "金额（元）": r['cost_per_ton']})
    
    st.dataframe(pd.DataFrame(fee_data), width="stretch", hide_index=True)

# ==================== 导出Excel ====================
if export_clicked and st.session_state.calc_results:
    wb = Workbook()
    ws = wb.active
    ws.title = "4措施"
    
    # 标题
    ws.merge_cells('A1:H1')
    ws['A1'] = project_name
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    row = 3
    
    # 机械措施
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "机械措施"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    row += 1
    
    headers = ['序号', '单体名称', '机械类型', '型号', '单位', '数量', '使用时间(天)', '备注']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    seq = 1
    for eq in st.session_state.equipment:
        ws.cell(row=row, column=1, value=seq)
        ws.cell(row=row, column=2, value=eq['building_idx'])
        ws.cell(row=row, column=3, value=eq['eq_type'].replace("/其他", ""))
        ws.cell(row=row, column=4, value=eq['eq_spec'])
        ws.cell(row=row, column=5, value="台")
        ws.cell(row=row, column=6, value=eq['eq_qty'])
        ws.cell(row=row, column=7, value=eq['required_days'])
        remark_parts = []
        if eq['remark']:
            remark_parts.append(eq['remark'])
        remark_parts.append(f"进出场{eq['mobilize_times']}次")
        ws.cell(row=row, column=8, value="; ".join(remark_parts))
        seq += 1
        row += 1
    
    # 辅助时间
    for aux in st.session_state.aux_times:
        if aux['work']:
            ws.cell(row=row, column=1, value=seq)
            ws.cell(row=row, column=2, value="")
            ws.cell(row=row, column=3, value="辅助时间")
            ws.cell(row=row, column=4, value=aux['work'])
            ws.cell(row=row, column=5, value="天")
            ws.cell(row=row, column=6, value="")
            ws.cell(row=row, column=7, value=aux['total'])
            ws.cell(row=row, column=8, value=f"{aux['eq_info']} {aux['remark']}".strip())
            seq += 1
            row += 1
    
    row += 1
    
    # 材料措施
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "材料措施"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    row += 1
    
    headers = ['序号', '单体名称', '措施内容', '单位', '数量', '使用时间(天)', '备注']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1
    
    seq = 1
    for mat in st.session_state.materials:
        ws.cell(row=row, column=1, value=seq)
        ws.cell(row=row, column=2, value=mat['building_idx'])
        ws.cell(row=row, column=3, value=mat['content'])
        ws.cell(row=row, column=4, value=mat['unit'])
        ws.cell(row=row, column=5, value=mat['qty'])
        ws.cell(row=row, column=6, value=mat['days'])
        ws.cell(row=row, column=7, value=mat['remark'])
        seq += 1
        row += 1
    
    row += 1
    
    # 备注
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "备注事项"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    row += 1
    
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = remarks
    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[row].height = 80
    row += 2
    
    # 签名栏
    ws[f'A{row}'] = f"编制人：{author}"
    ws[f'C{row}'] = f"审核人：{reviewer}"
    ws[f'E{row}'] = f"日期：{doc_date}"
    row += 1
    
    # 技术措施费
    if steel_weight > 0:
        ws[f'A{row}'] = f"技术措施费："
        ws[f'C{row}'] = f"用钢量：{steel_weight}吨"
    
    # 列宽
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 6
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 25
    
    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 提供下载
    file_name = f"{project_name}_机械措施表.xlsx" if project_name else "机械措施表.xlsx"
    st.download_button(
        label="📥 下载Excel文件",
        data=output,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width="stretch"
    )
